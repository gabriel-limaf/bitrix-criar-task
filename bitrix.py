import requests
import json
import PySimpleGUI as sg
import pandas as pd
from openpyxl import load_workbook
from time import sleep


def menu():  # Janela 1
    sg.theme('Dark Blue 3')
    layout = [[sg.Text('Bem-vindo(a) ao App Bitrix PM !!\n'
                       'O que deseja fazer?:\n')],
              [sg.Button('Criar tasks'), sg.Button('Campo CTI'), sg.Button('Cancelar')],
              [sg.Text('\nIndicium Tech - 2022', size=[75, 5], justification='center')]]
    return sg.Window('Menu', layout=layout, finalize=True, size=(500, 180))


def criar():  # Janela 2
    sg.theme('Dark Blue 3')
    layout = [
             [sg.Text('Caminho do Arquivo')],
             [sg.Input(), sg.FileBrowse(key='-SAIDA-', file_types=(('Text Files', '*.xls'),
                                                                   ('Text Files', '*.xlsx')))],
             [sg.Text('Informe seu ID do Bitrix')],
             [sg.InputText(key='bitrixID')],
             [sg.Text('Informe seu webhook do Bitrix')],
             [sg.InputText(key='bitrixKey')],
             [sg.Button('OK'), sg.Button('Voltar'), sg.Button('Cancelar')],
             [sg.Text('\nIndicium Tech - 2022', size=[75, 5], justification='center')]]
    return sg.Window('Bitrix - Criar tasks no Bitrix', layout=layout, finalize=True, size=(600, 300))


def sucesso():  # Janela 4
    sg.theme('DarkGreen')
    layout = [[sg.Text('Processo realizado com sucesso !')],
              [sg.Button('Voltar'), sg.Button('Cancelar')]]
    return sg.Window('SUCESSO', layout=layout, size=(300, 100), finalize=True)


def atualizar():  # Janela 5
    sg.theme('Dark Blue 3')
    layout = [[sg.Text('Caminho do Arquivo')],
              [sg.Input(), sg.FileBrowse(key='-SAIDA-', file_types=(('Text Files', '*.xls'),
                                                                    ('Text Files', '*.xlsx')))],
              [sg.Text('Informe seu ID do Bitrix')],
              [sg.InputText(key='bitrixID')],
              [sg.Text('Informe seu webhook do Bitrix')],
              [sg.InputText(key='bitrixKey')],
              [sg.Button('OK'), sg.Button('Voltar'), sg.Button('Cancelar')],
              [sg.Text('\nIndicium Tech - 2021', size=[75, 5], justification='center')]]
    return sg.Window('Bitrix - Lançar campo CTI no Bitrix', layout=layout, finalize=True, size=(600, 300))


janela1, janela2, janela4, janela5 = menu(), None, None, None
while True:
    window, event, values = sg.read_all_windows()
# Operações no MENU
    if window == janela1 and event == sg.WINDOW_CLOSED:
        break
    if window == janela1 and event == 'Cancelar':
        break
    if window == janela1 and event == 'Criar tasks':
        janela1.close()
        janela2 = criar()
    if window == janela1 and event == 'Campo CTI':
        janela1.close()
        janela5 = atualizar()

# Operações sucesso
    if window == janela4 and event == 'Voltar':
        janela4.close()
        janela1 = menu()
    if window == janela4 and event == 'Cancelar':
        break
    if window == janela4 and event == sg.WINDOW_CLOSED:
        break

# Operações escolher arquivo
    if window == janela2 and event == 'Voltar':
        janela2.close()
        janela1 = menu()
    if window == janela2 and event == 'Cancelar':
        break
    if window == janela2 and event == sg.WINDOW_CLOSED:
        break
    if window == janela5 and event == 'Voltar':
        janela5.close()
        janela1 = menu()
    if window == janela5 and event == 'Cancelar':
        break
    if window == janela5 and event == sg.WINDOW_CLOSED:
        break

# Criar task
    if window == janela2 and event == 'OK' and values['-SAIDA-'] != '':
        path_saida = values['-SAIDA-']
        bitrixID = values['bitrixID']
        bitrixKey = values['bitrixKey']
        df1 = pd.DataFrame(pd.read_excel(path_saida, sheet_name='Tarefas'))
        base_url = "https://indicium.bitrix24.com/rest/" + bitrixID + "/" + bitrixKey + "/"
        task_url_criar = "tasks.task.add"
        url_criar = base_url + task_url_criar
        task_url_checklist = "task.checklistitem.add"
        url_checklist = base_url + task_url_checklist
        for i, row in df1.iterrows():
            statusApi = str(row['status_api'])
            nomeTask = str(row['Task'])
            descricao = str(row['Descrição'])
            responsavel = str(row['Responsável'])
            deadline = str(row['Prazo final'])
            tempoEstimado = str(row['Horas estimadas'])
            criadoPor = str(row['Criada por'])
            participantes = str(row['Participantes']).split(';')
            observadores = str(row['Observadores']).split(';')
            projeto = str(row['Projeto'])
            marcadores = str(row['Marcadores']).split(';')
            campoCTI = str(row['CTI'])
            prioridade = str(row['Tarefa importante'])
            checklist = str(row['Lista de verificação']).split(';')
            if statusApi != 'Sucesso':
                if 'nan' not in marcadores:
                    payload = json.dumps({
                        "fields": {
                            "TITLE": nomeTask,
                            "DESCRIPTION": descricao,
                            "RESPONSIBLE_ID": responsavel,
                            "CREATED_BY": criadoPor,
                            "ACCOMPLICES": participantes,
                            "AUDITORS": observadores,
                            "DEADLINE": deadline,
                            "GROUP_ID": projeto,
                            "ALLOW_TIME_TRACKING": "Y",
                            "TIME_ESTIMATE": tempoEstimado,
                            "TAGS": marcadores,
                            "UF_AUTO_977208768718": campoCTI
                        }
                    })
                    headers = {
                        'Content-Type': 'application/json',
                    }
                    response = requests.request("POST", url_criar, headers=headers, data=payload)
                    print(response.text)
                    sleep(1)
                    obj = json.loads(response.text)
                    if response.status_code != 200:
                        df1.loc[i, 'status_api'] = 'Falha na criação da task pela API, ' \
                                                   'favor verificar se os campos na planilha estão corretos'
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
                        sleep(1)
                        continue
                    else:
                        idSaida = (obj['result']['task']['id'])
                        df1.loc[i, 'status_api'] = 'Sucesso'
                        df1.loc[i, 'ID'] = idSaida
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
# Criar checklist
                        if 'nan' not in checklist:
                            for x in checklist:
                                payload = json.dumps(
                                    [idSaida, {'TITLE': x, 'IS_COMPLETE': 'N'}]
                                )
                                headers = {
                                  'Content-Type': 'application/json',
                                }
                                response = requests.request("POST", url_checklist, headers=headers, data=payload)
                                print(response.text)
                                sleep(1)
# Criar task
                if 'nan' in marcadores:
                    payload = json.dumps({
                        "fields": {
                            "TITLE": nomeTask,
                            "DESCRIPTION": descricao,
                            "RESPONSIBLE_ID": responsavel,
                            "CREATED_BY": criadoPor,
                            "ACCOMPLICES": participantes,
                            "AUDITORS": observadores,
                            "DEADLINE": deadline,
                            "GROUP_ID": projeto,
                            "ALLOW_TIME_TRACKING": "Y",
                            "TIME_ESTIMATE": tempoEstimado,
                            "UF_AUTO_977208768718": campoCTI
                        }
                    })
                    headers = {
                        'Content-Type': 'application/json',
                    }
                    response = requests.request("POST", url_criar, headers=headers, data=payload)
                    print(response.text)
                    sleep(1)
                    obj = json.loads(response.text)
                    if response.status_code != 200:
                        df1.loc[i, 'status_api'] = 'Falha na criação da task pela API, ' \
                                                   'favor verificar se os campos na planilha estão corretos'
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
                        sleep(1)
                        continue
                    else:
                        idSaida = (obj['result']['task']['id'])
                        df1.loc[i, 'status_api'] = 'Sucesso'
                        df1.loc[i, 'ID'] = idSaida
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
# Criar checklist
                        if 'nan' not in checklist:
                            for x in checklist:
                                payload = json.dumps(
                                    [idSaida, {'TITLE': x, 'IS_COMPLETE': 'N'}]
                                )
                                headers = {
                                    'Content-Type': 'application/json',
                                }
                                response = requests.request("POST", url_checklist, headers=headers, data=payload)
                                print(response.text)
                                sleep(1)
        janela2.close()
        janela4 = sucesso()

# Atualizar campo CTI
    if window == janela5 and event == 'OK' and values['-SAIDA-'] != '':
        path_saida = values['-SAIDA-']
        bitrixID = values['bitrixID']
        bitrixKey = values['bitrixKey']
        df1 = pd.DataFrame(pd.read_excel(path_saida, sheet_name='Tarefas'))
        base_url = "https://indicium.bitrix24.com/rest/" + bitrixID + "/" + bitrixKey + "/"
        task_url = "tasks.task.update"
        url = base_url + task_url
        for i, row in df1.iterrows():
            idTask = str(row['ID'])
            campoCTI = str(row['CTI'])
            payload = json.dumps({
                "taskId": idTask,
                "fields": {
                    "UF_AUTO_977208768718": campoCTI
                }
            })
            headers = {
                'Content-Type': 'application/json',
            }
            response = requests.request("POST", url, headers=headers, data=payload)
            print(response.text)
            sleep(1)
        janela5.close()
        janela4 = sucesso()
